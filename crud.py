import openpyxl
from models import User, Movie
from typing import List, Optional
from fastapi import HTTPException

DB_PATH = "db.xlsx"
DELETED_PATH = "deleted.xlsx"

def __init_excel():
    try:
        wb = openpyxl.load_workbook(DB_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Users")
        wb.create_sheet("Movies")

        # Encabezados
        ws_users = wb["Users"]
        ws_users.append(["id", "name", "code", "id_deleted"])

        ws_movies = wb["Movies"]
        ws_movies.append(["id", "user_id", "title", "is_animated", "nominated", "nominated_year", "id_deleted"])

        wb.save(DB_PATH)

# Inicializa el archivo si no existe
__init_excel()

# -------------------------- USUARIOS --------------------------

def create_user(user: User):
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Users"]

    # Verificamos que no exista el ID ni el código
    for row in ws.iter_rows(min_row=2):
        if row[0].value == user.id or row[2].value == user.code:
            raise HTTPException(status_code=400, detail="ID o código ya existe")

    ws.append([user.id, user.name, user.code, False])
    wb.save(DB_PATH)

def get_user_by_id(user_id: int) -> Optional[User]:
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == user_id:
            return User(id=row[0].value, name=row[1].value, code=row[2].value, id_deleted=row[3].value)
    raise HTTPException(status_code=404, detail="Usuario no encontrado")

def get_user_by_code(code: str) -> User:
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2):
        if row[2].value == code:
            return User(id=row[0].value, name=row[1].value, code=row[2].value, id_deleted=row[3].value)
    raise HTTPException(status_code=404, detail="Usuario no encontrado")

def get_all_users() -> List[User]:
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Users"]
    return [
        User(id=row[0].value, name=row[1].value, code=row[2].value, id_deleted=row[3].value)
        for row in ws.iter_rows(min_row=2)
    ]

def get_active_users() -> List[User]:
    return [user for user in get_all_users() if not user.id_deleted]

def get_deleted_users() -> List[User]:
    return [user for user in get_all_users() if user.id_deleted]

def delete_user(user_id: int):
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == user_id:
            row[3].value = True
            wb.save(DB_PATH)
            return
    raise HTTPException(status_code=404, detail="Usuario no encontrado")

# -------------------------- PELÍCULAS --------------------------

def get_all_movies(include_deleted: bool = False) -> List[Movie]:
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Movies"]
    movies = []
    for row in ws.iter_rows(min_row=2):
        if not include_deleted and row[6].value:
            continue
        movies.append(Movie(
            id=row[0].value,
            user_id=row[1].value,
            title=row[2].value,
            is_animated=row[3].value,
            nominated=row[4].value,
            nominated_year=row[5].value,
            id_deleted=row[6].value,
        ))
    return movies

def create_movie(movie: Movie):
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Movies"]
    ws.append([
        movie.id,
        movie.user_id,
        movie.title,
        movie.is_animated,
        movie.nominated,
        movie.nominated_year,
        False,
    ])
    wb.save(DB_PATH)

def delete_movie(movie_id: int):
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Movies"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == movie_id:
            row[6].value = True
            save_deleted_movie(row)
            wb.save(DB_PATH)
            return
    raise HTTPException(status_code=404, detail="Película no encontrada")

def save_deleted_movie(row):
    try:
        wb = openpyxl.load_workbook(DELETED_PATH)
        ws = wb["Pelicula Eliminada"]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pelicula Eliminada"
        ws.append([
            "id", "user_id", "title", "is_animated", "nominated", "nominated_year", "id_deleted"
        ])
    ws.append([cell.value for cell in row])
    wb.save(DELETED_PATH)
